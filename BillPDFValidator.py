# -*- coding: utf-8 -*-
"""
Created on Fri Mar  8 10:38:10 2019

@author: RCHORGHA
"""
from openpyxl import Workbook
from tkinter import *
import PyPDF2
import re
import pandas as pd
#import tkinter as tk
from tkinter.ttk import *
#from tkinter import filedialog as fd
from tkinter import filedialog
from tkinter import messagebox as mb
import openpyxl
import os.path

# import openpyxl module 

global folder_path

def browse_button():
    # Allow user to select a directory and store it in global var
    # called folder_path
    filename = filedialog.askdirectory(title="Select path where bills & input sheet located")
    folder_path.set(filename)
    #print(filename)
def close_window(): 
    folder_path=None
    master.destroy()
    

def show_entry_fields():
        #Load data from excel on folder path
            print("Folder Path: %s\nShort: %s\nLong: %s\nProration: %s\nMRC: %s" % (folder_path.get(),var1.get(),var2.get(),var3.get(),var4.get()))
            if os.path.exists(folder_path.get()):
                data=pd.read_excel ('%s/PP_BAN.xlsx'%(folder_path.get()))
            else:
                mb.showinfo(title="Warning Message", message= "Select the correct folder path!")
                            
            #writedata=Workbook ('%s/PP_BAN.xlsx'%folder_path,'w')
            #ws=writedata.get_active_sheet
            #print(ws)
            #no of records
            rowcnt=data.iloc[:,0].count()
            print("No of records:%s"% int (rowcnt))
            #PP
            y = data.iloc[:, 0].values
            print (y[1])
            #BAN NO
            X = data.iloc[:, 1].values
            print (X[1])
            #Search String Short Description
            #z = data.iloc[:, 2].values
            #print (z[1])
            #Search String Short Description
            shortD = data.iloc[:, 2].values
            print (shortD[1])
            #Search String Long Description
            longD = data.iloc[:, 3].values
            print (longD[1])
            #Search String MRC
            MRC = data.iloc[:, 4].values
            print (MRC[1])
            #Search String Prorated MRC
            proratedMRC = data.iloc[:, 5].values
            print (proratedMRC[1])
                    
            wb = openpyxl.load_workbook('%s/PP_BAN.xlsx' % (folder_path.get())) 
            sheet = wb.active
            
    #try:
            #SHort Description validations
            if var1.get():
                for i in range (0,rowcnt):
                    # open the pdf file
                    filename=("%s/%s-%s.pdf"%(folder_path.get(),X[i],y[i]))
                    print("\nFilename:"+filename)
                    if os.path.isfile(filename):
                        pdfObject = PyPDF2.PdfFileReader("%s/%s-%s.pdf"%(folder_path.get(),X[i],y[i]))
                        print("this is ban " + str (X[i]))
                        print("this is PP " + str (y[i]))
                        # get number of pages
                        NumPages = pdfObject.getNumPages()
                        # define keyterms
                        String = str (shortD[i])
                        print(String)
                        #for j in range(0, NumPages):
                        #PageObj = pdfObject.getPage(2)
                        print("this is bill no " + str(i))
                        Text = PageObj.extractText()
                        #print(Text)
                        ResSearch = re.search(String, Text)
                        if (ResSearch!=None):
                            print("Short Description String Matched: "+str (ResSearch))
                            sheet.cell(row=i+2, column=7).value='Short Description Match Found'
                            wb.save('%s/PP_BAN.xlsx' % (folder_path.get()))
                        else:
                            print(Text)
                            sheet.cell(row=i+2, column=7).value='Short Description Match NOT Found'
                            wb.save('%s/PP_BAN.xlsx' % (folder_path.get())) 
                    else:
                        sheet.cell(row=i+2, column=7).value='No data found'
                        mb.showinfo(title="No such PDF file", message= "%s File name not found"%(filename))
                mb.showinfo(title="Sucess Message", message= "All bills verified with PP/SOC Short Description!")
                    
                #Long Description validations
            if var2.get():
                for i in range (0,rowcnt):
                    # open the pdf file
                    filename=("%s/%s-%s.pdf"%(folder_path.get(),X[i],y[i]))
                    print("\nFilename:"+filename)
                    if os.path.isfile(filename):
                        pdfObject = PyPDF2.PdfFileReader("%s/%s-%s.pdf"%(folder_path.get(),X[i],y[i]))
                        print("this is ban " + str (X[i]))
                        print("this is PP " + str (y[i]))
                        # get number of pages
                        NumPages = pdfObject.getNumPages()
                        # define keyterms
                        String = str (longD[i])
                        #String = String.replace(" ","")
                        print("Input String: "+String)
                        #for j in range(0, NumPages):
                        #PageObj = pdfObject.getPage(2)
                        print("this is bill no " + str(i))
                        #Text=PageObj.extractText().get_original_bytes()
                        #Text = PageObj.extractText().autodetect_pdfdocencoding()
                        Text = PageObj.extractText()
                        Text=Text.replace(" ","")
                        print(Text)
                        #ResSearch = re.search('\w*%s\w*'%String, Text)
                        ResSearch = Text.find(String)
                        if (ResSearch!=-1):
                            print("Long Description String Matched: "+str (ResSearch))
                            sheet.cell(row=i+2, column=8).value='Long Description Match Found'
                            wb.save('%s/PP_BAN.xlsx' % (folder_path.get()))
                        else:
                            print("Long Description String NOT Matched: "+str (ResSearch))
                            sheet.cell(row=i+2, column=8).value='Long Description Match NOT Found'
                            wb.save('%s/PP_BAN.xlsx' % (folder_path.get()))
                    else:
                        sheet.cell(row=i+2, column=8).value='No data found'
                        mb.showinfo(title="No such PDF file", message= "%s File name not found"%(filename))
                mb.showinfo(title="Sucess Message", message= "All bills verified with PP/SOC Long Description!")
                
            #MRC Description validations
            if var3.get():
                for i in range (0,rowcnt):
                    filename=("%s/%s-%s.pdf"%(folder_path.get(),X[i],y[i]))
                    print("\nFilename:"+filename)
                    if os.path.isfile(filename):
                        # open the pdf file
                        pdfObject = PyPDF2.PdfFileReader("%s/%s-%s.pdf"%(folder_path.get(),X[i],y[i]))
                        print("this is ban " + str (X[i]))
                        print("this is PP " + str (y[i]))
                        # get number of pages
                        NumPages = pdfObject.getNumPages()
                        # define keyterms
                        String = str (MRC[i])
                        print(String)
                        #for j in range(0, NumPages):
                        #PageObj = pdfObject.getPage(2)
                        print("this is bill no " + str(i))
                        Text = PageObj.extractText()
                        #print(Text)
                        ResSearch = re.search(String, Text)
                        if (ResSearch!=None):
                            print("MRC String Matched: "+str (ResSearch))
                            sheet.cell(row=i+2, column=9).value='MRC Match Found'
                            wb.save('%s/PP_BAN.xlsx' % (folder_path.get()))
                        else:
                            sheet.cell(row=i+2, column=9).value='MRC Match NOT Found'
                            wb.save('%s/PP_BAN.xlsx' % (folder_path.get()))
                    else:
                        sheet.cell(row=i+2, column=9).value='No data found'
                        mb.showinfo(title="No such PDF file", message= "%s File name not found"%(filename))
                mb.showinfo(title="Sucess Message", message= "All bills verified with PP/SOC MRC!")

            #Prorated MRC validations            
            if var4.get():
                for i in range (0,rowcnt):
                    filename=("%s/%s-%s.pdf"%(folder_path.get(),X[i],y[i]))
                    print("\nFilename:"+filename)
                    if os.path.isfile(filename):
                    # open the pdf file
                        pdfObject = PyPDF2.PdfFileReader("%s/%s-%s.pdf"%(folder_path.get(),X[i],y[i]))
                        print("this is ban " + str (X[i]))
                        print("this is PP " + str (y[i]))
                        # get number of pages
                        NumPages = pdfObject.getNumPages()
                        # define keyterms
                        String = str (proratedMRC[i])
                        print(String)
                        #for j in range(0, NumPages):
                        #PageObj = pdfObject.getPage(2)
                        print("this is bill no " + str(i))
                        Text = PageObj.extractText()
                        #print(Text)
                        ResSearch = re.search(String, Text)
                        if (ResSearch!=None):
                            print("Proration MRC String Matched: "+str (ResSearch))
                            sheet.cell(row=i+2, column=10).value='Prorated MRC Match Found'
                            wb.save('%s/PP_BAN.xlsx' % (folder_path.get()))
                        else:
                            sheet.cell(row=i+2, column=10).value='Prorated MRC Match NOT Found'
                            wb.save('%s/PP_BAN.xlsx' % (folder_path.get()))
                    else:
                        sheet.cell(row=i+2, column=10).value='No data found'
                        mb.showinfo(title="No such PDF file", message= "%s File name not found"%(filename))
                mb.showinfo(title="Sucess Message", message= "All bills verified with PP/SOC Prorated MRC!")

master = Tk()
master.geometry("400x200+500+300")
master.title("Bill PDF Validation Tool")

Label(master, text="Choose Folder Location").grid(row=0, column=0)
folder_path = StringVar()
brwsBtn = Button(master, text="Browse", command=browse_button).grid (row=0, column=1)
var1 = IntVar()
Checkbutton(master, text="Price Plan/SOC Short Description", variable=var1).grid(row=2, sticky=W)
var2 = IntVar()
Checkbutton(master, text="Price Plan/SOC Long Description", variable=var2).grid(row=3, sticky=W)
var3 = IntVar()
Checkbutton(master, text="Price Plan/SOC Full MRC", variable=var3).grid(row=4, sticky=W)
var4 = IntVar()
Checkbutton(master, text="Price Plan/SOC Prorated MRC", variable=var4).grid(row=5, sticky=W)
#e1 = Entry(master)
#e2 = Entry(master)
Button(master, text='Cancel', command=close_window).grid(row=7, column=0, sticky=W, pady=3)
Button(master, text='Close', command=close_window).grid(row=7, column=1, sticky=W, pady=3)
Button(master, text='Search', command=show_entry_fields).grid(row=7, column=2, sticky=W, pady=3)

mainloop()
#except IOError:
#mb.showerror(title="Error Messsage", message="Please make sure you provided all required inputs and closed worksheet")
            

    
        
